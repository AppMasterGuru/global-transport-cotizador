"""
GT Cotizador — Seed reference data from SharePoint Excel files.

Downloads three files from JP's OneDrive and seeds the local SQLite DB:

  1. DATA COLOADERS.xlsx  → providers table
     (MSL, CRAFT, SACO, VANGUARD, ECU WORLDWIDE — contact directory)

  2. LISTA CRÉDITOS.xlsx  → credit_registry table
     (CLIENTES LOCALES, AGENTES INTERNACIONALES, PROVEEDORES DE SERVICIO)

Run from the cotizador/ directory:
    python scripts/seed_reference_data.py

Safe to re-run — skips existing entries (company+contact_name+email for providers,
company+category for credit_registry).
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))

from dotenv import load_dotenv
load_dotenv(_ROOT / ".env")

import io
import openpyxl
import requests

from core.db import init_db, seed_credit_registry
from core.drive import get_graph_token, GRAPH_BASE, _DRIVE_ID, download_excel
from core.providers import seed_providers

# ── File IDs (confirmed 2026-05-15) ──────────────────────────────────────────
_COLOADERS_FILE_ID    = "01QZQ7OQFQ7NKQAWYVPBH3RXKTS7X54AGS"
_LISTA_CREDITOS_FILE_ID = "01QZQ7OQGLPXK7J6A2YFHKA24ZWPM75SXC"

COLOADERS_FILE_ID    = os.getenv("COLOADERS_FILE_ID",    _COLOADERS_FILE_ID)
LISTA_CREDITOS_FILE_ID = os.getenv("LISTA_CREDITOS_FILE_ID", _LISTA_CREDITOS_FILE_ID)


# ══════════════════════════════════════════════════════════════════════════════
# Parsers
# ══════════════════════════════════════════════════════════════════════════════

def parse_coloaders(xlsx_bytes: bytes) -> list[dict]:
    """
    Parse DATA COLOADERS.xlsx → list of provider contact dicts.

    Sheet structure (Hoja1):
      CONSOLIDADOR (NVOCC) | CONTACTO | CARGO | EMAIL | TELEFONO
    Company name is in column A; blank means inherit from previous row.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    contacts: list[dict] = []
    current_company = ""

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue
        v0 = str(row[0]).strip() if row[0] else ""
        v1 = str(row[1]).strip() if row[1] else ""
        v2 = str(row[2]).strip() if row[2] else ""
        v3 = str(row[3]).strip() if row[3] else ""
        v4 = str(row[4]).strip() if row[4] else ""

        # Skip header row
        if v0.upper() == "CONSOLIDADOR (NVOCC)":
            continue

        # Update current company when column A has a value
        if v0:
            current_company = v0

        if not current_company or not v1:
            continue

        # Skip rows that look like decorations (single letter, etc.)
        if len(v1) <= 1:
            continue

        contacts.append({
            "company":      current_company,
            "contact_name": v1,
            "role":         v2,
            "email":        v3,
            "phone":        v4,
        })

    wb.close()
    return contacts


def parse_lista_creditos(xlsx_bytes: bytes) -> list[dict]:
    """
    Parse LISTA CRÉDITOS.xlsx → list of credit registry entry dicts.

    Sheet mapping:
      CLIENTES LOCALES         → category='local_client'
      AGENTES INTERNACIONALES  → category='international_agent'
      PROVEEDORES DE SERVICIO  → category='service_provider'

    Columns (CLIENTES LOCALES):   EMPRESA | DÍAS | CONDICIÓN
    Columns (AGENTES INTER.):     EMPRESA | PAÍS | DÍAS | CONDICIÓN
    Columns (PROVEEDORES):        EMPRESA | DÍAS | CONDICIÓN | LÍNEA | COMENTARIO
    """
    _SHEET_CATEGORY = {
        "CLIENTES LOCALES":        "local_client",
        "AGENTES INTERNACIONALES": "international_agent",
        "PROVEEDORES DE SERVICIO": "service_provider",
    }

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    entries: list[dict] = []

    for sheet_name in wb.sheetnames:
        category = _SHEET_CATEGORY.get(sheet_name.strip().upper())
        if not category:
            continue
        ws = wb[sheet_name]

        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            v0 = str(row[0]).strip() if row[0] else ""
            if not v0 or v0.upper() in ("EMPRESA", "CONSOLIDADOR"):
                continue

            if category == "international_agent":
                # EMPRESA | PAÍS | DÍAS | CONDICIÓN | [credit_line]
                country     = str(row[1]).strip() if row[1] else ""
                credit_days = int(row[2]) if row[2] is not None and str(row[2]).isdigit() else None
                condition   = str(row[3]).strip() if row[3] else ""
                credit_line = int(row[4]) if row[4] is not None and str(row[4]).lstrip('-').isdigit() else None
                entries.append({
                    "company": v0, "category": category, "country": country,
                    "credit_days": credit_days, "condition": condition,
                    "credit_line": credit_line, "notes": "",
                })
            elif category == "service_provider":
                # EMPRESA | DÍAS | CONDICIÓN | LÍNEA | COMENTARIO
                credit_days = int(row[1]) if row[1] is not None else None
                condition   = str(row[2]).strip() if row[2] else ""
                credit_line = int(row[3]) if row[3] is not None and str(row[3]).lstrip('-').isdigit() else None
                notes       = str(row[4]).strip() if row[4] else ""
                entries.append({
                    "company": v0, "category": category, "country": "",
                    "credit_days": credit_days, "condition": condition,
                    "credit_line": credit_line, "notes": notes,
                })
            else:
                # local_client: EMPRESA | DÍAS | CONDICIÓN
                credit_days = int(row[1]) if row[1] is not None else None
                condition   = str(row[2]).strip() if row[2] else ""
                entries.append({
                    "company": v0, "category": category, "country": "Peru",
                    "credit_days": credit_days, "condition": condition,
                    "credit_line": None, "notes": "",
                })

    wb.close()
    return entries


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("=" * 60)
    print("GT Cotizador — Seed Reference Data from SharePoint")
    print("=" * 60)

    # Ensure DB schema exists
    init_db()
    print("✅ DB schema verified")

    token = get_graph_token()
    if not token:
        print("❌ No Graph token — check credentials in .env")
        sys.exit(1)

    # ── 1. DATA COLOADERS → providers ────────────────────────────────────────
    print(f"\n📥 Downloading DATA COLOADERS.xlsx (id={COLOADERS_FILE_ID})…")
    raw = download_excel(COLOADERS_FILE_ID)
    print(f"   {len(raw):,} bytes received")

    contacts = parse_coloaders(raw)
    print(f"   Parsed {len(contacts)} contacts")
    for c in contacts:
        print(f"     {c['company']:<25} {c['contact_name']:<30} {c['role']}")

    inserted_p = seed_providers(contacts)
    print(f"✅ providers: {inserted_p} new rows inserted (skipped existing)")

    # ── 2. LISTA CRÉDITOS → credit_registry ──────────────────────────────────
    print(f"\n📥 Downloading LISTA CRÉDITOS.xlsx (id={LISTA_CREDITOS_FILE_ID})…")
    raw2 = download_excel(LISTA_CREDITOS_FILE_ID)
    print(f"   {len(raw2):,} bytes received")

    entries = parse_lista_creditos(raw2)
    print(f"   Parsed {len(entries)} credit registry entries")

    inserted_c = seed_credit_registry(entries)
    print(f"✅ credit_registry: {inserted_c} new rows inserted (skipped existing)")

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(f"DONE")
    print(f"  providers:       {inserted_p} new contacts")
    print(f"  credit_registry: {inserted_c} new entries")
    print("=" * 60)


if __name__ == "__main__":
    main()
