"""
Microsoft Graph / SharePoint rate card reader.
Reads GT's TARIFAS folder from JP's personal OneDrive on SharePoint.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
AUTH SETUP — barney.elliott@gmail.com as external guest user
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Step 1: Register an Azure AD app (one-time, takes 5 min)
  • Go to: https://portal.azure.com → Azure Active Directory → App registrations
  • New registration:
      Name: "GT Cotizador"
      Supported account types: "Accounts in any organizational directory and personal Microsoft accounts"
      Redirect URI: leave blank (we use device code flow)
  • After creation, note the Application (client) ID → set as GRAPH_CLIENT_ID in .env
  • Under API permissions → Add a permission → Microsoft Graph → Delegated:
      Files.Read.All, Sites.Read.All, offline_access
  • Grant admin consent if prompted (JP or Vania must approve)

Step 2: Get initial token via device code flow (run once from terminal)
  python3 -c "
  import requests, json, os
  client_id = '<your GRAPH_CLIENT_ID>'
  tenant    = 'globaltransportsac285.onmicrosoft.com'

  # Request device code
  r = requests.post(
      f'https://login.microsoftonline.com/{tenant}/oauth2/v2.0/devicecode',
      data={'client_id': client_id, 'scope': 'Files.Read.All Sites.Read.All offline_access'},
  )
  code_data = r.json()
  print(code_data['message'])  # visit URL, enter code shown

  # Poll for token (after you've authenticated in the browser)
  input('Press Enter once you have authenticated in the browser...')
  r2 = requests.post(
      f'https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token',
      data={
          'client_id': client_id,
          'grant_type': 'urn:ietf:params:oauth:grant-type:device_code',
          'device_code': code_data['device_code'],
      },
  )
  tokens = r2.json()
  print('ACCESS_TOKEN:', tokens.get('access_token','')[:40], '...')
  print('REFRESH_TOKEN:', tokens.get('refresh_token','')[:40], '...')
  "

Step 3: Get Drive ID and Folder ID (run once after token works)
  python3 -c "
  import requests
  headers = {'Authorization': 'Bearer <your_access_token>'}
  # Get JP's drive ID
  r = requests.get('https://graph.microsoft.com/v1.0/drives', headers=headers)
  print(r.json())
  # Then navigate to TARIFAS folder to get its ID
  drive_id = '<drive_id from above>'
  r2 = requests.get(
      f'https://graph.microsoft.com/v1.0/drives/{drive_id}/root:'
      '/Documents/GLOBAL TRANSPORT - COMERCIAL/TARIFAS',
      headers=headers,
  )
  print('TARIFAS folder ID:', r2.json().get('id'))
  "

Step 4: Fill .env
  GRAPH_CLIENT_ID=<from Step 1>
  GRAPH_TENANT_ID=globaltransportsac285.onmicrosoft.com
  GRAPH_ACCESS_TOKEN=<from Step 2>
  GRAPH_REFRESH_TOKEN=<from Step 2>
  SHAREPOINT_DRIVE_ID=<from Step 3>
  TARIFAS_FOLDER_ID=<from Step 3>

Note: access tokens expire in ~1 hour. This module auto-refreshes using
the refresh token (valid ~90 days). After 90 days, repeat Step 2.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

FUTURE INTEGRATION — ISO 9001 Procedures (Google Drive)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GT's ISO 9001 quality procedures are stored in a Google Drive folder:
  https://drive.google.com/drive/folders/17uWrMaZXXF-mDuXOgZKmaSvX2Xx-sS_S

This folder is NOT yet integrated. Integration requires:
  1. Google Drive API credentials (service account or OAuth 2.0)
  2. GOOGLE_DRIVE_CLIENT_ID / GOOGLE_DRIVE_CLIENT_SECRET in .env
  3. A new reader function (similar to get_rate_card) to fetch procedure docs

Target use cases:
  - Surface relevant SOP when a quote involves restricted cargo
  - Attach compliance checklist to APPROVED quotes for BASC audit trail
  - Cross-check customs agent / carrier selection against approved vendor list

Environment variable: ISO_GOOGLE_DRIVE_URL (set in .env, not yet consumed by code)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

from __future__ import annotations

import io
import json
import os
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import requests

from core.db import audit

# ── Config from .env ──────────────────────────────────────────────────────────

GRAPH_BASE       = "https://graph.microsoft.com/v1.0"
_CLIENT_ID       = os.getenv("GRAPH_CLIENT_ID", "")
_CLIENT_SECRET   = os.getenv("GRAPH_CLIENT_SECRET", "")
_TENANT          = os.getenv("GRAPH_TENANT_ID", "globaltransportsac285.onmicrosoft.com")
_ACCESS_TOKEN    = os.getenv("GRAPH_ACCESS_TOKEN", "")
_REFRESH_TOKEN   = os.getenv("GRAPH_REFRESH_TOKEN", "")
_DRIVE_ID        = os.getenv("SHAREPOINT_DRIVE_ID", "")
_TARIFAS_FOLDER_ID   = os.getenv("TARIFAS_FOLDER_ID", "")
# Handling fees file (HANDLING AEREO.xlsx) — confirmed 2026-05-15
_HANDLING_AEREO_FILE_ID = os.getenv("HANDLING_AEREO_FILE_ID", "01QZQ7OQFULNIELXYXEFB2HQ3GULHKIE2V")

# Two supported auth modes:
#   1. Delegated (device code):   GRAPH_ACCESS_TOKEN + GRAPH_REFRESH_TOKEN
#   2. Client credentials (app):  GRAPH_CLIENT_ID + GRAPH_CLIENT_SECRET
#      — confirmed working 2026-05-15 via scripts/get_drive_ids.py
_HAS_DELEGATED   = bool(_ACCESS_TOKEN and _DRIVE_ID)
_HAS_APP_AUTH    = bool(_CLIENT_ID and _CLIENT_SECRET and _DRIVE_ID)
_CONFIGURED      = _HAS_DELEGATED or _HAS_APP_AUTH

# ── Cache ─────────────────────────────────────────────────────────────────────

CACHE_DIR      = Path(os.getenv("CACHE_DIR", "/tmp/gt_cotizador_cache"))
CACHE_TTL_SECS = int(os.getenv("RATE_CARD_CACHE_TTL_HOURS", "24")) * 3600

# ── Rate card file name patterns ──────────────────────────────────────────────

# Actual filenames in TARIFAS (confirmed 2026-05-15 via scripts/get_drive_ids.py):
#   TARIFARIO AGENTES EXPO GT - 2025 - V2.xlsx
#   TARIFARIO AGENTES IMPO GT - 2025 - V2.xlsx
_FILE_PATTERNS = {
    "expo": re.compile(r"TARIFARIO\s+AGENTES\s+EXPO\s+GT", re.IGNORECASE),
    "impo": re.compile(r"TARIFARIO\s+AGENTES\s+IMPO\s+GT", re.IGNORECASE),
}

# ── Expected sheets per direction ─────────────────────────────────────────────

EXPO_SHEETS = [
    "FCL EXW EXPO", "FCL FOB EXPO",
    "LCL EXW EXPO", "LCL FOB EXPO",
    "AIR EXW EXPO", "AIR FCA EXPO",
]
IMPO_SHEETS = [
    "FCL DDP IMPO", "FCL DAP IMPO",
    "LCL DDP IMPO", "LCL DAP IMPO",
    "AIR DAP IMPO",
]

# ── Mock fallback rates (used when Graph not configured) ──────────────────────
# These mirror the structure returned by get_rate_cards() so callers
# don't need to branch on live vs mock.

_MOCK_RATES: dict[str, dict] = {
    "expo": {
        "FCL FOB EXPO": {
            "carrier": "Mock Naviera",
            "port_of_loading": "Callao, Peru",
            "port_of_destination": "TBD",
            "transit_time": "TBD",
            "frequency": "Weekly",
            "validity": "Mock data — load real rate cards",
            "cost_lines": [],
            "sell_lines": [],
            "flete_internacional_usd": 0.0,
            "source": "mock",
        },
        "LCL FOB EXPO": {
            "carrier": "MSL (mock)",
            "port_of_loading": "Callao, Peru",
            "port_of_destination": "TBD",
            "transit_time": "TBD",
            "frequency": "Weekly",
            "validity": "Mock data — load real rate cards",
            "cost_lines": [],
            "sell_lines": [],
            "flete_internacional_usd": 0.0,
            "source": "mock",
        },
        "AIR EXW EXPO": {
            "carrier": "LAN (mock)",
            "port_of_loading": "Lima LIM",
            "port_of_destination": "TBD",
            "transit_time": "TBD",
            "frequency": "Daily",
            "validity": "Mock data — load real rate cards",
            "cost_lines": [],
            "sell_lines": [],
            "flete_internacional_usd": 0.0,
            "source": "mock",
        },
    },
    "impo": {
        "FCL DAP IMPO": {
            "carrier": "Mock Naviera",
            "port_of_loading": "TBD",
            "port_of_destination": "Callao, Peru",
            "transit_time": "TBD",
            "frequency": "Weekly",
            "validity": "Mock data — load real rate cards",
            "cost_lines": [],
            "sell_lines": [],
            "flete_internacional_usd": 0.0,
            "source": "mock",
        },
        "LCL DAP IMPO": {
            "carrier": "MSL (mock)",
            "port_of_loading": "TBD",
            "port_of_destination": "Callao, Peru",
            "transit_time": "TBD",
            "frequency": "Weekly",
            "validity": "Mock data — load real rate cards",
            "cost_lines": [],
            "sell_lines": [],
            "flete_internacional_usd": 0.0,
            "source": "mock",
        },
        "AIR DAP IMPO": {
            "carrier": "LAN (mock)",
            "port_of_loading": "TBD",
            "port_of_destination": "Lima LIM",
            "transit_time": "TBD",
            "frequency": "Daily",
            "validity": "Mock data — load real rate cards",
            "cost_lines": [],
            "sell_lines": [],
            "flete_internacional_usd": 0.0,
            "source": "mock",
        },
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# Token management
# ══════════════════════════════════════════════════════════════════════════════

# Runtime token store — starts from .env, updated by refresh
_token_store: dict[str, str] = {
    "access_token":  _ACCESS_TOKEN,
    "refresh_token": _REFRESH_TOKEN,
}


def get_graph_token() -> str:
    """
    Return a valid Graph access token.

    Priority:
      1. In-memory token (previously fetched this process lifetime).
      2. Client credentials grant (GRAPH_CLIENT_ID + GRAPH_CLIENT_SECRET).
         Tokens expire in ~1h; auto-refreshed on 401 via _refresh_access_token().
      3. Delegated refresh token (GRAPH_REFRESH_TOKEN), if present.
      4. Returns empty string if not configured.
    """
    current = _token_store.get("access_token", "")
    if current:
        return current
    # No token in store yet — try to acquire one
    if _HAS_APP_AUTH:
        _refresh_access_token()  # populates _token_store via client credentials
    return _token_store.get("access_token", "")


def _refresh_access_token() -> bool:
    """
    Acquire or refresh an access token. Updates _token_store in place.
    Tries client credentials first (preferred — no expiry management needed
    beyond the 1h token lifetime). Falls back to refresh token if available.
    Returns True on success, False on failure.
    """
    # ── Path A: client credentials (app auth) ─────────────────────────────────
    if _CLIENT_ID and _CLIENT_SECRET:
        try:
            resp = requests.post(
                f"https://login.microsoftonline.com/{_TENANT}/oauth2/v2.0/token",
                data={
                    "client_id":     _CLIENT_ID,
                    "client_secret": _CLIENT_SECRET,
                    "scope":         "https://graph.microsoft.com/.default",
                    "grant_type":    "client_credentials",
                },
                timeout=10,
            )
            if resp.status_code == 200:
                _token_store["access_token"] = resp.json()["access_token"]
                return True
        except requests.RequestException:
            pass

    # ── Path B: delegated refresh token ───────────────────────────────────────
    refresh = _token_store.get("refresh_token", "") or _REFRESH_TOKEN
    if not (refresh and _CLIENT_ID):
        return False
    try:
        resp = requests.post(
            f"https://login.microsoftonline.com/{_TENANT}/oauth2/v2.0/token",
            data={
                "client_id":    _CLIENT_ID,
                "grant_type":   "refresh_token",
                "refresh_token": refresh,
                "scope":        "Files.Read.All Sites.Read.All offline_access",
            },
            timeout=10,
        )
        if resp.status_code == 200:
            tokens = resp.json()
            _token_store["access_token"]  = tokens["access_token"]
            _token_store["refresh_token"] = tokens.get("refresh_token", refresh)
            return True
    except requests.RequestException:
        pass
    return False


def _headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {get_graph_token()}",
        "Accept":        "application/json",
    }


def _graph_get(url: str) -> requests.Response:
    """
    GET a Graph API URL. On 401, attempts one token refresh and retries.
    Raises requests.HTTPError on persistent failure.
    """
    resp = requests.get(url, headers=_headers(), timeout=15)
    if resp.status_code == 401:
        if _refresh_access_token():
            resp = requests.get(url, headers=_headers(), timeout=15)
    resp.raise_for_status()
    return resp


# ══════════════════════════════════════════════════════════════════════════════
# Graph API — file operations
# ══════════════════════════════════════════════════════════════════════════════

def list_tarifas_folder() -> list[dict]:
    """
    List files inside the TARIFAS folder on JP's OneDrive.
    Returns list of Graph DriveItem dicts.
    Returns [] if not configured or on any error.
    """
    if not _CONFIGURED:
        return []
    try:
        if _TARIFAS_FOLDER_ID:
            url = f"{GRAPH_BASE}/drives/{_DRIVE_ID}/items/{_TARIFAS_FOLDER_ID}/children"
        else:
            # Fallback: navigate by path
            url = (
                f"{GRAPH_BASE}/drives/{_DRIVE_ID}/root:"
                "/Documents/GLOBAL TRANSPORT - COMERCIAL/TARIFAS:/children"
            )
        resp = _graph_get(url)
        return resp.json().get("value", [])
    except (requests.RequestException, KeyError, ValueError):
        return []


def _get_file_metadata(file_id: str) -> dict[str, Any]:
    """
    Return Graph DriveItem metadata for a single file (includes lastModifiedDateTime).
    Returns {} on any error.
    """
    try:
        url = f"{GRAPH_BASE}/drives/{_DRIVE_ID}/items/{file_id}"
        resp = _graph_get(url)
        return resp.json()
    except (requests.RequestException, ValueError):
        return {}


def download_excel(file_id: str) -> bytes:
    """
    Download an Excel file from JP's OneDrive by its Graph item ID.
    Returns raw bytes. Raises requests.RequestException on failure.
    """
    url = f"{GRAPH_BASE}/drives/{_DRIVE_ID}/items/{file_id}/content"
    # Download endpoint returns a redirect; requests follows it automatically
    resp = requests.get(url, headers=_headers(), timeout=60, allow_redirects=True)
    if resp.status_code == 401:
        if _refresh_access_token():
            resp = requests.get(url, headers=_headers(), timeout=60, allow_redirects=True)
    resp.raise_for_status()
    return resp.content


# ══════════════════════════════════════════════════════════════════════════════
# Cache
# ══════════════════════════════════════════════════════════════════════════════

def _cache_key(file_name: str, last_modified: str) -> str:
    """Stable cache key: filename slug + last-modified timestamp."""
    slug = re.sub(r"[^a-z0-9]", "_", file_name.lower())[:60]
    ts_slug = re.sub(r"[^0-9]", "", last_modified)[:14]
    return f"ratecard_{slug}_{ts_slug}"


def _cache_path(key: str) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    return CACHE_DIR / f"{key}.json"


def _cache_load(key: str) -> dict | None:
    """Return cached dict if it exists and is < CACHE_TTL_SECS old, else None."""
    p = _cache_path(key)
    if not p.exists():
        return None
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        age = time.time() - data.get("_cached_at", 0)
        if age < CACHE_TTL_SECS:
            return data
    except (json.JSONDecodeError, OSError):
        pass
    return None


def _cache_save(key: str, data: dict) -> None:
    """Write data to cache with current timestamp."""
    try:
        payload = {**data, "_cached_at": time.time()}
        _cache_path(key).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError:
        pass


# ══════════════════════════════════════════════════════════════════════════════
# Excel parser
# ══════════════════════════════════════════════════════════════════════════════

# Keywords for header extraction — matches against cell string values.
# Covers both FCL layout ("STEAMSHIP LINE", "FRECUENCY", "TT / DAYS") and
# LCL layout ("COLOADER", "PORT OF DEPARTURE", "TRANSIT TIME", "FREQUENCY").
_HEADER_KEYS = {
    "port_of_loading":     re.compile(r"port\s+of\s+(load|depart)", re.I),
    "port_of_destination": re.compile(r"port\s+of\s+dest|destino|destination", re.I),
    "carrier":             re.compile(r"carrier|naviera|aerolinea|airline|coloader|steamship", re.I),
    "transit_time":        re.compile(r"transit\s*time|tiempo\s+transito|tránsito|\btt\s*/", re.I),
    "frequency":           re.compile(r"frequen|frecuen", re.I),
    "validity":            re.compile(r"validity|validez|vigencia", re.I),
}

# Keywords to identify cost line rows
_COST_SECTION_RE  = re.compile(r"^cost\b|^costo|^costos", re.I)
_TARIFA_SECTION_RE = re.compile(r"tarifa\s*neta|sell\s*side|tarifa\s+venta", re.I)
_TARIFA_FIJA_RE   = re.compile(r"tarifa\s*fija|fixed\s*rate", re.I)
_TOTAL_RE         = re.compile(r"^total", re.I)
_SKIP_RE          = re.compile(r"^(concept|concepto|description|item|#)$", re.I)

# Cost type markers (values in the TYPE column)
_VALID_TYPES      = {"BL", "CONT", "TN", "M3", "KG", "VOL", "CBM", "SHIP", "DOC"}


def _cell_str(cell) -> str:
    """Return the stripped string value of a cell, or ''."""
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def _cell_float(cell) -> float | None:
    """Return float value or None."""
    v = cell.value
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _scan_header(ws) -> dict[str, str]:
    """
    Scan the worksheet for header key-value pairs.
    Strategy: find a cell matching a header keyword, take value from same row
    in the next non-empty column to the right (or the cell to the right).
    """
    header: dict[str, str] = {k: "" for k in _HEADER_KEYS}

    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            text = _cell_str(cell)
            if not text:
                continue
            for field, pattern in _HEADER_KEYS.items():
                if header[field]:           # already found
                    continue
                if pattern.search(text):
                    # Look for the value in cells to the right on the same row
                    for j in range(i + 1, min(i + 6, len(row))):
                        val = _cell_str(row[j])
                        if val:
                            header[field] = val
                            break
    return header


def _find_table_header(ws) -> tuple[int | None, dict, dict]:
    """
    Find the rate-card table header row and discover column indices for
    the left (cost) and right (sell/tarifa neta) tables.

    Both FCL and LCL sheets have a header row with CONCEPT/CONCEPTO in col B (idx 1)
    and a matching right-side table in a later column group.

    Returns: (header_row_0based_idx, left_col_map, right_col_map)
    Each col_map: {concept, cost, usd, igv, type}  (igv/type may be absent)
    """
    _CONCEPT_RE = re.compile(r"^concep", re.I)    # concept or concepto
    _COST_RE    = re.compile(r"^costo?$", re.I)   # cost or costo
    _USD_RE     = re.compile(r"^usd$", re.I)
    _IGV_RE     = re.compile(r"^igv$", re.I)
    _TYPE_RE    = re.compile(r"^type$|^tn[/\\]m3$|^type\b", re.I)

    for row_idx, row in enumerate(ws.iter_rows()):
        cells = list(row)
        # Find all columns that contain CONCEPT/CONCEPTO
        concept_cols = [i for i, c in enumerate(cells) if _CONCEPT_RE.match(_cell_str(c))]
        if not concept_cols:
            continue

        tables = []
        for start in concept_cols:
            t: dict[str, int] = {"concept": start}
            for j in range(start + 1, min(start + 8, len(cells))):
                v = _cell_str(cells[j])
                if _COST_RE.match(v):
                    t.setdefault("cost", j)
                elif _IGV_RE.match(v):
                    t.setdefault("igv", j)
                elif _USD_RE.match(v):
                    t.setdefault("usd", j)
                elif _TYPE_RE.match(v):
                    t.setdefault("type", j)
            if "cost" in t and "usd" in t:
                tables.append(t)

        if tables:
            left  = tables[0]
            right = tables[1] if len(tables) > 1 else {}
            return row_idx, left, right

    return None, {}, {}


def _scan_cost_table(ws) -> tuple[list[dict], list[dict]]:
    """
    Scan for cost (left) and sell/tarifa neta (right) line items.

    Discovers column positions dynamically from the header row so it works
    for both FCL (5 left cols) and LCL/AIR (4 left cols) sheet layouts.

    Returns:
        cost_lines  — list of {concept, cost_usd, type, notes, fixed, is_total}
        sell_lines  — list of {concept, cost_usd, igv_usd, total_usd, fixed, is_total}
    """
    header_row_idx, left_cols, right_cols = _find_table_header(ws)
    if header_row_idx is None or not left_cols:
        return [], []

    cost_lines: list[dict] = []
    sell_lines: list[dict] = []

    all_rows = list(ws.iter_rows())

    for row in all_rows[header_row_idx + 1:]:
        cells = list(row)
        if not cells:
            continue

        row_text = " ".join(_cell_str(c) for c in cells if _cell_str(c))
        if not row_text:
            continue
        # Stop at notes / annotations section
        if re.match(r"notes?:", row_text, re.I):
            break

        is_fixed = bool(_TARIFA_FIJA_RE.search(row_text))

        # ── Left (cost) table ─────────────────────────────────────────────────
        lc  = left_cols.get("concept", -1)
        lu  = left_cols.get("usd", -1)
        lt  = left_cols.get("type", -1)

        concept = _cell_str(cells[lc]) if 0 <= lc < len(cells) else ""

        # TOTAL may appear offset from the concept column (some sheets put it
        # one col to the right of the concept column)
        is_total = bool(_TOTAL_RE.match(concept))
        if not concept:
            search_range = range(max(0, lc - 1), min(lc + 5, len(cells)))
            for k in search_range:
                if _TOTAL_RE.match(_cell_str(cells[k])):
                    concept = "TOTAL"
                    is_total = True
                    break

        if concept and not _SKIP_RE.match(concept):
            usd_val  = _cell_float(cells[lu]) if 0 <= lu < len(cells) else None
            type_val = _cell_str(cells[lt])   if 0 <= lt < len(cells) else ""
            if usd_val is not None:
                cost_lines.append({
                    "concept":  concept,
                    "cost_usd": usd_val,
                    "type":     type_val,
                    "notes":    "",
                    "fixed":    is_fixed,
                    "is_total": is_total,
                })

        # ── Right (sell/tarifa neta) table ────────────────────────────────────
        if not right_cols:
            continue

        rc    = right_cols.get("concept", -1)
        ru    = right_cols.get("usd", -1)
        ri    = right_cols.get("igv", -1)
        rcost = right_cols.get("cost", -1)

        r_concept = _cell_str(cells[rc]) if 0 <= rc < len(cells) else ""
        is_total_r = bool(_TOTAL_RE.match(r_concept))
        if not r_concept:
            search_range = range(max(0, rc - 1), min(rc + 5, len(cells)))
            for k in search_range:
                if _TOTAL_RE.match(_cell_str(cells[k])):
                    r_concept = "TOTAL"
                    is_total_r = True
                    break

        if r_concept and not _SKIP_RE.match(r_concept):
            usd_r  = _cell_float(cells[ru])    if 0 <= ru    < len(cells) else None
            igv_r  = _cell_float(cells[ri])    if 0 <= ri    < len(cells) else None
            cost_r = _cell_float(cells[rcost]) if 0 <= rcost < len(cells) else None
            if usd_r is not None or cost_r is not None:
                sell_lines.append({
                    "concept":   r_concept,
                    "cost_usd":  cost_r or 0.0,
                    "igv_usd":   igv_r  or 0.0,
                    "total_usd": usd_r if usd_r is not None else (cost_r or 0.0),
                    "fixed":     is_fixed,
                    "is_total":  is_total_r,
                })

    return cost_lines, sell_lines


def _total_cost(cost_lines: list[dict]) -> float:
    """Sum all cost_usd values that aren't sub-totals."""
    return round(sum(l["cost_usd"] for l in cost_lines if not l.get("is_total")), 2)


def _total_sell(sell_lines: list[dict]) -> float:
    """Return the explicit TOTAL row, or sum if absent."""
    totals = [l["total_usd"] for l in sell_lines if l.get("is_total")]
    if totals:
        return round(totals[-1], 2)
    return round(sum(l["total_usd"] for l in sell_lines if not l.get("is_total")), 2)


def _parse_sheet(ws, sheet_name: str) -> dict:
    """
    Parse a single rate-card worksheet.
    Returns a structured dict regardless of whether parsing succeeded fully.
    """
    header = _scan_header(ws)
    cost_lines, sell_lines = _scan_cost_table(ws)

    # Best-effort: derive flete from first cost line if header not extracted
    flete = 0.0
    for line in cost_lines:
        concept_lower = line["concept"].lower()
        if any(k in concept_lower for k in
               ["ocean freight", "air freight", "flete", "freight"]):
            flete = line["cost_usd"]
            break
    if not flete and cost_lines:
        flete = cost_lines[0]["cost_usd"]

    # Determine mode/incoterm from sheet name
    parts = sheet_name.upper().split()
    mode = parts[0] if parts else "UNK"          # FCL / LCL / AIR
    incoterm = parts[1] if len(parts) > 1 else "" # EXW / FOB / DDP / etc.
    direction = parts[2] if len(parts) > 2 else ""

    return {
        "sheet_name":            sheet_name,
        "mode":                  mode,
        "incoterm":              incoterm,
        "direction":             direction,
        "carrier":               header.get("carrier", ""),
        "port_of_loading":       header.get("port_of_loading", ""),
        "port_of_destination":   header.get("port_of_destination", ""),
        "transit_time":          header.get("transit_time", ""),
        "frequency":             header.get("frequency", ""),
        "validity":              header.get("validity", ""),
        "cost_lines":            cost_lines,
        "sell_lines":            sell_lines,
        "flete_internacional_usd": flete,
        "total_cost_usd":        _total_cost(cost_lines),
        "total_sell_usd":        _total_sell(sell_lines),
        "source":                "live",
    }


def _parse_excel_bytes(xlsx_bytes: bytes) -> dict[str, dict]:
    """
    Parse all sheets in an Excel workbook.
    Returns {sheet_name: parsed_dict}.
    Skips blank or unreadable sheets silently.
    """
    results: dict[str, dict] = {}
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes),
            data_only=True,   # read computed values, not formulas
            read_only=True,
        )
        for name in wb.sheetnames:
            try:
                ws = wb[name]
                results[name] = _parse_sheet(ws, name)
            except Exception:
                results[name] = {"sheet_name": name, "parse_error": True, "source": "live"}
        wb.close()
    except Exception as exc:
        results["__error__"] = {"error": str(exc)}
    return results


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════

def _find_file_id(direction: str) -> tuple[str | None, str | None, str | None]:
    """
    Search the TARIFAS folder for the expo or impo rate card.
    Returns (file_id, file_name, last_modified) or (None, None, None).
    """
    pattern = _FILE_PATTERNS.get(direction.lower())
    if not pattern:
        return None, None, None

    items = list_tarifas_folder()
    for item in items:
        name = item.get("name", "")
        if pattern.search(name):
            return (
                item.get("id"),
                name,
                item.get("lastModifiedDateTime", ""),
            )
    return None, None, None


def get_rate_cards(direction: str) -> dict[str, dict]:
    """
    Download and parse the rate card Excel for 'expo' or 'impo'.

    Caches for CACHE_TTL_SECS (default 24h) keyed on file name + last_modified.
    Falls back to mock data if Graph is not configured or on any error.
    Logs RATE_CARD_FETCHED (or RATE_CARD_FALLBACK) to audit trail.

    Returns: {sheet_name: parsed_sheet_dict}
    """
    direction = direction.lower()
    if direction not in ("expo", "impo"):
        raise ValueError(f"direction must be 'expo' or 'impo', got {direction!r}")

    if not _CONFIGURED:
        audit("RATE_CARD_FALLBACK", None, "system", {
            "direction": direction,
            "reason": "Graph not configured — using mock rates",
        })
        return _MOCK_RATES.get(direction, {})

    try:
        file_id, file_name, last_modified = _find_file_id(direction)

        if not file_id:
            audit("RATE_CARD_FALLBACK", None, "system", {
                "direction": direction,
                "reason": f"Rate card file not found in TARIFAS folder",
            })
            return _MOCK_RATES.get(direction, {})

        # Check cache
        key = _cache_key(file_name or direction, last_modified or "")
        cached = _cache_load(key)
        if cached:
            return {k: v for k, v in cached.items() if not k.startswith("_")}

        # Cache miss — download fresh
        xlsx_bytes = download_excel(file_id)
        sheets = _parse_excel_bytes(xlsx_bytes)

        _cache_save(key, sheets)

        audit("RATE_CARD_FETCHED", None, "system", {
            "direction":     direction,
            "file_name":     file_name,
            "last_modified": last_modified,
            "sheet_count":   len(sheets),
            "cached_key":    key,
        })
        return sheets

    except Exception as exc:
        audit("RATE_CARD_FALLBACK", None, "system", {
            "direction": direction,
            "reason":    f"Error: {exc}",
        })
        return _MOCK_RATES.get(direction, {})


def get_live_rates(mode: str, incoterm: str, direction: str) -> dict:
    """
    Return the best matching rate sheet for this mode + incoterm + direction.

    Falls back to mock rates if Graph is not configured or sheet not found.
    This is the integration point for core/transport.py callers.

    Args:
        mode:      "fcl", "lcl", "aereo"
        incoterm:  "FOB", "CIF", "EXW", "FCA", "DAP", "DDP", etc.
        direction: "expo" (outbound from Peru) or "impo" (inbound to Peru)

    Returns dict with at minimum:
        {
            carrier, port_of_loading, port_of_destination,
            transit_time, frequency, validity,
            flete_internacional_usd, cost_lines, sell_lines,
            source: "live" | "mock",
        }
    """
    mode_upper     = mode.upper()
    incoterm_upper = incoterm.upper()

    try:
        sheets = get_rate_cards(direction)
    except Exception:
        sheets = _MOCK_RATES.get(direction.lower(), {})

    # Target sheet name: e.g. "FCL FOB EXPO", "LCL DAP IMPO", "AIR FCA EXPO"
    dir_upper   = direction.upper()
    target_name = f"{mode_upper} {incoterm_upper} {dir_upper}"

    if target_name in sheets:
        return sheets[target_name]

    # Fallback 1: match on mode only (e.g. any FCL EXPO sheet)
    for name, sheet in sheets.items():
        if name.startswith(f"{mode_upper} ") and dir_upper in name:
            return {**sheet, "source": sheet.get("source", "live"), "_matched": "mode_only"}

    # Fallback 2: return mock for this mode
    mock_dir = _MOCK_RATES.get(direction.lower(), {})
    for name, mock in mock_dir.items():
        if name.startswith(f"{mode_upper} "):
            return mock

    # Ultimate fallback: empty structure with source=mock
    return {
        "carrier": "",
        "port_of_loading": "",
        "port_of_destination": "",
        "transit_time": "TBD",
        "frequency": "TBD",
        "validity": "Mock data",
        "cost_lines": [],
        "sell_lines": [],
        "flete_internacional_usd": 0.0,
        "source": "mock",
    }


# ══════════════════════════════════════════════════════════════════════════════
# HANDLING AEREO — air ground-handling fees per airline
# Source: HANDLING AEREO.xlsx on JP's OneDrive (confirmed 2026-05-15)
# Sheets: TALMA, SHOHIN, SAASA
# Columns: Aerolínea | Valor Vta. (USD) | IGV | Monto Total | Counter
# ══════════════════════════════════════════════════════════════════════════════

def _parse_handling_aereo(xlsx_bytes: bytes) -> list[dict]:
    """
    Parse HANDLING AEREO.xlsx.
    Returns list of:
      {airline, handler (sheet name), counter, net_usd, igv_usd, total_usd}
    Skips header rows and empty rows.
    """
    fees: list[dict] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_seen = False
            for row in ws.iter_rows(values_only=True):
                if not any(v is not None for v in row):
                    continue
                v0 = str(row[0]).strip() if row[0] is not None else ""
                if not header_seen:
                    # Header row: first cell contains "Aerolínea" or "AEROL"
                    if "aerol" in v0.lower():
                        header_seen = True
                    continue
                if not v0:
                    continue
                try:
                    net_usd   = float(row[1]) if row[1] is not None else 0.0
                    igv_usd   = float(row[2]) if row[2] is not None else 0.0
                    total_usd = float(row[3]) if row[3] is not None else net_usd * 1.18
                    counter   = str(row[4]).strip() if len(row) > 4 and row[4] else sheet_name
                    fees.append({
                        "airline":   v0,
                        "handler":   sheet_name,
                        "counter":   counter,
                        "net_usd":   round(net_usd, 4),
                        "igv_usd":   round(igv_usd, 4),
                        "total_usd": round(total_usd, 4),
                    })
                except (TypeError, ValueError, IndexError):
                    continue
        wb.close()
    except Exception:
        pass
    return fees


def get_air_handling_fees() -> list[dict]:
    """
    Download and parse HANDLING AEREO.xlsx from JP's OneDrive.
    Returns list of fee dicts. Cached for CACHE_TTL_SECS (default 24h).
    Returns [] when Graph is not configured or file unavailable.
    """
    if not _CONFIGURED or not _HANDLING_AEREO_FILE_ID:
        return []
    try:
        meta     = _get_file_metadata(_HANDLING_AEREO_FILE_ID)
        last_mod = meta.get("lastModifiedDateTime", "unknown")
        key      = _cache_key("handling_aereo", last_mod)
        cached   = _cache_load(key)
        if cached:
            return cached.get("fees", [])

        xlsx_bytes = download_excel(_HANDLING_AEREO_FILE_ID)
        fees       = _parse_handling_aereo(xlsx_bytes)
        _cache_save(key, {"fees": fees})
        audit("HANDLING_AEREO_FETCHED", None, "system", {
            "file_id":    _HANDLING_AEREO_FILE_ID,
            "fee_count":  len(fees),
            "last_mod":   last_mod,
        })
        return fees
    except Exception as exc:
        audit("HANDLING_AEREO_ERROR", None, "system", {"error": str(exc)})
        return []


def get_air_handling_fee(carrier: str) -> dict | None:
    """
    Look up the handling fee record for a given airline/carrier name.

    Matching priority:
      1. Exact match (case-insensitive)
      2. Counter/short-name exact match (e.g. "LAN" matches "LAN AIRLINES / LAN PERU")
      3. Token match — carrier words appear in airline name or vice versa

    Returns fee dict or None if no match found.
    """
    if not carrier:
        return None
    fees = get_air_handling_fees()
    if not fees:
        return None

    needle = carrier.strip().upper()

    # Pass 1: exact airline name
    for f in fees:
        if f["airline"].upper() == needle:
            return f

    # Pass 2: counter match
    for f in fees:
        if f["counter"].upper() == needle:
            return f

    # Pass 3: token containment — every word in needle appears in airline name
    needle_tokens = set(needle.split())
    for f in fees:
        airline_upper = f["airline"].upper()
        if needle_tokens and needle_tokens.issubset(set(airline_upper.split())):
            return f
        # Or: needle is a substring of the airline name
        if len(needle) >= 3 and needle in airline_upper:
            return f

    return None
