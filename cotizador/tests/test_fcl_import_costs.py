"""
Tests for FCL import local costs by naviera — Abel Parte 2 (2026-06-19).

Abel's explicit written rule: FCL import charges are THC + ISPS + MBL
emission ONLY. Parsed (not hardcoded) from:
  - G. LOCALES sheet -> THCD (20'/40') + ISPS/adicional (20'/40')
  - EMISION MBL sheet -> MBL emission cost per naviera

HOLD (TODO abel-Q3): there is a second import structure (EXPO_IMPO
IMPORTACIÓN sheet's VB IMPORTACION layer) that may stack with or replace
the above — explicitly NOT wired into the import total this session.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from core.fcl_import_costs import (
    build_vb_importacion_totals,
    get_fcl_import_local_costs,
    get_g_locales_data,
    get_mbl_data,
    get_vb_importacion_data,
    parse_g_locales_sheet,
    parse_import_vb_sheet,
    parse_mbl_sheet,
    parse_vb_importacion_sheet,
)


def _make_g_locales_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "G. LOCALES"
    rows = [
        ["NAVIERA", "AGENTE \nMARITIMO", "POD", "THCD", None, "ADICIONALES", None, None],
        [None, None, None, "20'", "40'", "CONCEPTO", "20'", "40'"],
        ["CMA CGM / APL", "IAN TAYLOR", "Callao", 65, 70, "ISPS", 14, 14],
        [None, None, None, 40, 65, None, None, None],
        ["COSCO / OOCL", "COSCO PERU", "Callao", 55, 55, "ISPS", 6, 6],
        ["MAERSK / SEALAND", "COLUMBUS", "Callao", 110, 110, "DOC FEE", 55, 55],
        ["MSC", "MSC PERU", "Callao", 65, 65, None, None, None],
        ["SEABOARD", "CITIKOLD", "Callao", "NO COBRA", "NO COBRA", None, None, None],
        ["HAMBURG SUD / ALIANCA", "COLUMBUS", "Callao", 90, 90, "ISPS",
         "USD 16.00 \nEUR 13.00", "USD 16.00 \nEUR 13.00"],
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_mbl_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EMISION MBL"
    rows = [
        ["NAVIERA", "COSTO POR EMISION MBL", "SEA WAYBILL / TELEX"],
        ["APL", "USD 55.00 + IGV", "NO"],
        ["CMA CGM", "USD 55.00 + IGV", "SI", None, "VB FCL VÍA GT"],
        ["COSCO / OOCL", "USD 30.00 + IGV", "SI"],
        ["MAERSK / SEALAND", "DOC FEE USD 55.00", "SI"],
        ["MSC", "USD 57.00 + IGV", "NO"],
        ["SEABOARD", "USD 35.00 + IGV", "NO"],
        ["ONE", "USD 29,50 + IGV", "SI"],
        ["HAMBURG SUD", "DOC FEE USD 30.93", "NO"],
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def g_locales():
    wb = openpyxl.load_workbook(io.BytesIO(_make_g_locales_xlsx()), data_only=True, read_only=True)
    return parse_g_locales_sheet(wb["G. LOCALES"])


@pytest.fixture
def mbl():
    wb = openpyxl.load_workbook(io.BytesIO(_make_mbl_xlsx()), data_only=True, read_only=True)
    return parse_mbl_sheet(wb["EMISION MBL"])


def _make_import_vb_xlsx() -> bytes:
    """
    Three VB IMPORTACIÓN blocks, replicating the real layout/values from
    EXPO_IMPO.xlsx's IMPORTACIÓN sheet (Client Data/Part 2_Abel/):
      - "BOX FEE - EXPO MSK" / "COVERAGE FEE - EXPO MSK" desglose ->
        naviera-identifiable (MSK token) -> MAERSK / SEALAND, total 150.50.
      - "CMA - COORDINACIÓN Y SUPERVISIÓN DE EMBARQUE" desglose ->
        naviera-identifiable (CMA token) -> CMA CGM / APL, total 194.75.
      - "DESPACHO DEL CONTENEDOR" / "DESPACHO DOCUMENTARIO" desglose -> no
        naviera token -> unidentified, total 294 (not naviera-attributed,
        same as the unidentified blocks Abel's sheet leaves ambiguous).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IMPORTACIÓN"
    rows = [
        ["CALLAO"],
        ["VISTO BUENO (BOX FEE | COVERAGE FEE)", 150.5, 64.5, 215],
        ["DESGLOSE DEL VISTO BUENO (IMPORTACIÓN)"],
        ["BOX FEE - EXPO MSK", 101.5, 43.5, 145, "CONTENEDOR"],
        ["COVERAGE FEE - EXPO MSK", 49, 21, 70, "CONTENEDOR"],
        ["DEMARES", "GATE IN | 20' & 40'", 237, 42.66, 279.66, "CONTENEDOR"],
        ["CALLAO"],
        ["VISTO BUENO (COORDINACIÓN Y SUPERVISIÓN DE EMBARQUE | AGENCY FEE)", 194.75, 35.055, 229.805],
        ["DESGLOSE DEL VISTO BUENO (IMPORTACIÓN)"],
        ["CMA - COORDINACIÓN Y SUPERVISIÓN DE EMBARQUE", 190, 34.2, 224.2, "CONTENEDOR"],
        ["IMUPESA", "GATE IN | 20' & 40'", 205, 36.9, 241.9, "CONTENEDOR"],
        ["CALLAO"],
        ["VISTO BUENO (DESPACHO DE CONTENEDOR | DESPACHO DOCUMENTARIO)", 294, 52.92, 346.92],
        ["DESGLOSE DEL VISTO BUENO (IMPORTACIÓN)"],
        ["DESPACHO DEL CONTENEDOR", 174, 31.32, 205.32, "CONTENEDOR"],
        ["DESPACHO DOCUMENTARIO", 120, 21.6, 141.6, "BL"],
        ["MEDLOG", "GATE IN | 20' &40", 215, 38.7, 253.7, "CONTENEDOR"],
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def import_vb():
    wb = openpyxl.load_workbook(io.BytesIO(_make_import_vb_xlsx()), data_only=True, read_only=True)
    return parse_import_vb_sheet(wb["IMPORTACIÓN"])


class TestGLocalesParsing:
    def test_cma_apl_thc(self, g_locales):
        e = g_locales["CMA CGM / APL"]
        assert e["thc_20"] == pytest.approx(65.0, rel=0.001)
        assert e["thc_40"] == pytest.approx(70.0, rel=0.001)

    def test_cma_apl_isps(self, g_locales):
        e = g_locales["CMA CGM / APL"]
        assert e["adicional_concept"] == "ISPS"
        assert e["adicional_20"] == pytest.approx(14.0, rel=0.001)
        assert e["adicional_40"] == pytest.approx(14.0, rel=0.001)

    def test_cosco_oocl(self, g_locales):
        e = g_locales["COSCO / OOCL"]
        assert e["thc_20"] == pytest.approx(55.0, rel=0.001)
        assert e["thc_40"] == pytest.approx(55.0, rel=0.001)
        assert e["adicional_20"] == pytest.approx(6.0, rel=0.001)
        assert e["adicional_40"] == pytest.approx(6.0, rel=0.001)

    def test_maersk_sealand_doc_fee(self, g_locales):
        e = g_locales["MAERSK / SEALAND"]
        assert e["thc_20"] == pytest.approx(110.0, rel=0.001)
        assert e["thc_40"] == pytest.approx(110.0, rel=0.001)
        assert e["adicional_concept"] == "DOC FEE"
        assert e["adicional_20"] == pytest.approx(55.0, rel=0.001)
        assert e["adicional_40"] == pytest.approx(55.0, rel=0.001)

    def test_seaboard_no_cobra(self, g_locales):
        e = g_locales["SEABOARD"]
        assert e["thc_20"] is None
        assert e["thc_40"] is None
        assert e["no_cobra"] is True


class TestMblParsing:
    def test_apl_55_plus_igv(self, mbl):
        e = mbl["APL"]
        assert e["amount"] == pytest.approx(55.0, rel=0.001)
        assert e["currency"] == "USD"
        assert e["plus_igv"] is True

    def test_cosco_oocl_30_plus_igv(self, mbl):
        e = mbl["COSCO / OOCL"]
        assert e["amount"] == pytest.approx(30.0, rel=0.001)
        assert e["currency"] == "USD"

    def test_maersk_sealand_doc_fee_55(self, mbl):
        e = mbl["MAERSK / SEALAND"]
        assert e["amount"] == pytest.approx(55.0, rel=0.001)
        assert "DOC FEE" in e["raw"]

    def test_comma_decimal_parsed(self, mbl):
        # "USD 29,50 + IGV" -> 29.50, not 2950
        e = mbl["ONE"]
        assert e["amount"] == pytest.approx(29.50, rel=0.001)


class TestFclImportLocalCostsCombined:
    """Abel's rule: THC + ISPS + MBL only."""

    def test_cma_apl_combined(self, g_locales, mbl):
        # ISPS corrected to 39 by Abel Q4 2026-06-19 (was 14 from the raw
        # sheet) — see TestCmaAplIgvExemptOverrideQ4 for the full override.
        c = get_fcl_import_local_costs(g_locales, mbl, "CMA CGM / APL")
        assert c["thc_20"] == pytest.approx(65.0, rel=0.001)
        assert c["isps_20"] == pytest.approx(39.0, rel=0.001)
        assert c["mbl_usd"] == pytest.approx(55.0, rel=0.001)

    def test_cosco_oocl_combined(self, g_locales, mbl):
        c = get_fcl_import_local_costs(g_locales, mbl, "COSCO / OOCL")
        assert c["thc_20"] == pytest.approx(55.0, rel=0.001)
        assert c["isps_20"] == pytest.approx(6.0, rel=0.001)
        assert c["mbl_usd"] == pytest.approx(30.0, rel=0.001)

    def test_maersk_sealand_combined(self, g_locales, mbl):
        c = get_fcl_import_local_costs(g_locales, mbl, "MAERSK / SEALAND")
        assert c["thc_20"] == pytest.approx(110.0, rel=0.001)
        assert c["mbl_usd"] == pytest.approx(55.0, rel=0.001)

    def test_seaboard_no_cobra_handled_cleanly(self, g_locales, mbl):
        c = get_fcl_import_local_costs(g_locales, mbl, "SEABOARD")
        assert c["thc_20"] is None
        assert c["thc_no_cobra"] is True
        assert c["mbl_usd"] == pytest.approx(35.0, rel=0.001)

    def test_unknown_naviera_returns_none(self, g_locales, mbl):
        assert get_fcl_import_local_costs(g_locales, mbl, "NOT-A-REAL-LINE") is None


class TestCmaAplIgvExemptOverrideQ4:
    """Abel Parte 2 Q4 (2026-06-19): CMA CGM / APL THC and ISPS are
    corrected (65 / 39) and IGV-exempt — overrides the raw G. LOCALES
    sheet figures (which would otherwise default to IGV-applicable)."""

    def test_thc_is_65_no_igv(self, g_locales, mbl):
        c = get_fcl_import_local_costs(g_locales, mbl, "CMA CGM / APL")
        assert c["thc_20"] == pytest.approx(65.0, rel=0.001)
        assert c["thc_40"] == pytest.approx(65.0, rel=0.001)
        assert c["thc_igv_applicable"] is False

    def test_isps_is_39_no_igv(self, g_locales, mbl):
        c = get_fcl_import_local_costs(g_locales, mbl, "CMA CGM / APL")
        assert c["isps_20"] == pytest.approx(39.0, rel=0.001)
        assert c["isps_40"] == pytest.approx(39.0, rel=0.001)
        assert c["isps_igv_applicable"] is False

    def test_other_navieras_default_igv_applicable_true(self, g_locales, mbl):
        c = get_fcl_import_local_costs(g_locales, mbl, "COSCO / OOCL")
        assert c["thc_igv_applicable"] is True
        assert c["isps_igv_applicable"] is True


class TestHamburgSudInactiveQ5:
    """Abel confirmed no cargo with this carrier June 19 (Q5) — Hamburg Sud /
    Alianca is removed from the active naviera list even though it still
    has rows in the source sheets."""

    def test_hamburg_sud_returns_none(self, g_locales, mbl):
        assert get_fcl_import_local_costs(g_locales, mbl, "HAMBURG SUD / ALIANCA") is None


class TestImportVbSheetParsing:
    """Abel Parte 2 Q3 (closed 2026-06-20): the EXPO_IMPO IMPORTACIÓN
    sheet's VB IMPORTACION layer stacks with THC+ISPS+MBL rather than
    replacing it. Only naviera-identifiable blocks (explicit token in the
    desglose concept text) get attributed — same no-guessing rule already
    applied to the export VISTO BUENO blocks in fcl_naviera_costs.py."""

    def test_maersk_identified_via_msk_token(self, import_vb):
        e = import_vb["MAERSK / SEALAND"]
        assert e["vb_importacion_usd"] == pytest.approx(150.5, rel=0.001)

    def test_cma_identified_via_cma_token(self, import_vb):
        e = import_vb["CMA CGM / APL"]
        assert e["vb_importacion_usd"] == pytest.approx(194.75, rel=0.001)

    def test_unidentified_block_not_attributed(self, import_vb):
        # The 294-total block (DESPACHO DEL CONTENEDOR / DESPACHO
        # DOCUMENTARIO) has no naviera token in its desglose, so it must
        # NOT be guessed into any naviera key.
        assert "MSC" not in import_vb
        assert len(import_vb) == 2


class TestFclImportLocalCostsVbStackingQ3:
    """get_fcl_import_local_costs sums the VB-import layer with
    THC+ISPS+MBL when a naviera-identified vb_importacion entry exists."""

    def test_maersk_combined_includes_vb_importacion(self, g_locales, mbl, import_vb):
        c = get_fcl_import_local_costs(g_locales, mbl, "MAERSK / SEALAND", vb_importacion=import_vb)
        assert c["thc_20"] == pytest.approx(110.0, rel=0.001)
        assert c["mbl_usd"] == pytest.approx(55.0, rel=0.001)
        assert c["vb_importacion_usd"] == pytest.approx(150.5, rel=0.001)

    def test_naviera_without_vb_entry_is_none(self, g_locales, mbl, import_vb):
        # COSCO / OOCL has no naviera-identified VB import block.
        c = get_fcl_import_local_costs(g_locales, mbl, "COSCO / OOCL", vb_importacion=import_vb)
        assert c["vb_importacion_usd"] is None

    def test_no_vb_importacion_arg_defaults_to_none(self, g_locales, mbl):
        # Backward-compatible default — no vb_importacion dict supplied.
        c = get_fcl_import_local_costs(g_locales, mbl, "CMA CGM / APL")
        assert c["vb_importacion_usd"] is None


def _make_vb_importacion_xlsx() -> bytes:
    """
    Subset of the real "VB IMPORTACION" sheet in Gastos de Importacion en
    Callao por Naviera.xlsx (Client Data/Part 2_Abel/) — every block is
    explicitly keyed by naviera name in column A, unlike the EXPO_IMPO
    IMPORTACIÓN sheet's VISTO BUENO blocks (no naviera guessing needed).
    Columns: NAVIERA, CONCEPTO, POD, UNIT, P. UNIT, MONEDA, FACTURA.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VB IMPORTACION"
    rows = [
        ["NAVIERA", "CONCEPTO", "POD", "UNIT", "P. UNIT", "MONEDA", "FACTURA"],
        ["CMA-CGM / APL", "COORDINACIÓN Y SUPERVISIÓN DE DESCARGA", "Callao", "Cntr.", 190, "USD", "IAN TAYLOR"],
        [None, "ADMINISTRACIÓN Y PROTECCIÓN DE EQUIPOS", None, "Cntr.", 35, "USD", "IAN TAYLOR"],
        [None, "GASTOS ADMINISTRATIVOS (2.5% FACTURA)", None, "Factura", 0, "USD", "IAN TAYLOR"],
        [None, "GATE IN", None, "Cntr.", 187, "USD", "UNIMAR"],
        ["COSCO / OOCL", "Servicio de Administración de Contenedores (SAC)", "Callao", "Cntr.", 120, "PEN", "COSCO"],
        [None, "VoBo 1 Cntr.", None, "Cntr.", 540, "PEN", "COSCO"],
        [None, "VoBo 2 a 5 Cntr.", None, "Cntr.", 470, "PEN", "COSCO"],
        [None, "VoBo 6 a más Cntr.", None, "Cntr.", 390, "PEN", "COSCO"],
        [None, "Gate in", None, "Cntr.", 192.5, "USD", "TPP"],
        ["ONE", "Gastos Administrativos", "Callao", "Factura", 10, "USD", "MERCATOR"],
        [None, "Box fee", None, "Cntr.", 155, "USD", "MERCATOR"],
        [None, "SCAC", None, "Cntr.", 42, "USD", "MERCATOR"],
        [None, "Doc Fee", None, "BL", 115, "USD", "MERCATOR"],
        [None, "Gate In", None, "Cntr.", 198, "USD", "CONTRANS / NEPTUNIA"],
        ["SEABOARD", "CONTAINER DELIVERY ORDER FEE", "Callao", "Cntr.", 116, "USD", "CITIKOLD"],
        [None, "TEXT PERCENTAGE ROW (ROBUSTNESS)", None, "BL", "2.5% NO NUMERIC VALUE", None, "TEST"],
        [None, "GASTOS ADMINISTRATIVOS", None, "Factura", 30, "USD", "CITIKOLD"],
        [None, "DOC FEE", None, "BL", 140.98, "USD", "CITIKOLD"],
        ["Montos no incluyen IGV"],
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def vb_importacion_raw():
    wb = openpyxl.load_workbook(io.BytesIO(_make_vb_importacion_xlsx()), data_only=True, read_only=True)
    return parse_vb_importacion_sheet(wb["VB IMPORTACION"])


class TestVbImportacionSheetParsing:
    """Gastos VB IMPORTACION sheet — every block explicitly keyed by naviera
    name, resolving ABEL_FOLLOWUPS.md item #2's unattributed-block problem
    outright (no naviera guessing needed for any naviera on this sheet)."""

    def test_all_navieras_attributed(self, vb_importacion_raw):
        assert set(vb_importacion_raw) == {"CMA CGM / APL", "COSCO / OOCL", "ONE", "SEABOARD"}

    def test_hyphenated_naviera_name_normalized(self, vb_importacion_raw):
        # Sheet header is "CMA-CGM / APL" (hyphen); canonical key used
        # elsewhere in this module (G. LOCALES) is "CMA CGM / APL" (space).
        assert "CMA-CGM / APL" not in vb_importacion_raw
        assert "CMA CGM / APL" in vb_importacion_raw

    def test_gate_in_rows_excluded_from_concepts(self, vb_importacion_raw):
        concepts = [c["concept"].upper() for c in vb_importacion_raw["CMA CGM / APL"]["concepts"]]
        assert not any("GATE IN" in c for c in concepts)

    def test_gate_in_case_variant_excluded(self, vb_importacion_raw):
        concepts = [c["concept"].upper() for c in vb_importacion_raw["COSCO / OOCL"]["concepts"]]
        assert not any("GATE IN" in c for c in concepts)

    def test_non_numeric_p_unit_skipped(self, vb_importacion_raw):
        concepts = [c["concept"] for c in vb_importacion_raw["SEABOARD"]["concepts"]]
        assert "TEXT PERCENTAGE ROW (ROBUSTNESS)" not in concepts

    def test_pen_currency_preserved_raw(self, vb_importacion_raw):
        sac = next(c for c in vb_importacion_raw["COSCO / OOCL"]["concepts"] if "SAC" in c["concept"].upper())
        assert sac["currency"] == "PEN"
        assert sac["p_unit"] == pytest.approx(120.0, rel=0.001)

    def test_footer_note_not_treated_as_naviera(self, vb_importacion_raw):
        assert "MONTOS NO INCLUYEN IGV" not in vb_importacion_raw


class TestBuildVbImportacionTotals:
    """build_vb_importacion_totals() converts the raw per-concept parse into
    a flat {naviera: {vb_importacion_usd}} dict — same shape consumed by
    get_fcl_import_local_costs(vb_importacion=...), so it slots in as a
    drop-in replacement for the EXPO_IMPO-sourced parse_import_vb_sheet()
    without changing that function's signature.

    TODO(abel-F1F4): figures here come from Abel's own Gastos de Importacion
    en Callao por Naviera.xlsx "VB IMPORTACION" sheet, not yet validated
    against a real quote — confirm via Abel's F1-F4 scenario run, same
    validation gate as the rest of FCL.
    """

    def test_cma_apl_flat_per_container(self, vb_importacion_raw):
        # 190 + 35 + 0 = 225, x1 container, USD only — no PEN conversion needed.
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=1, exchange_rate=3.75)
        assert totals["CMA CGM / APL"]["vb_importacion_usd"] == pytest.approx(225.0, rel=0.001)

    def test_cma_apl_cntr_unit_scales_with_container_count(self, vb_importacion_raw):
        # Cntr.-unit concepts (190 + 35) scale x2; Factura concept (0) does not.
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=2, exchange_rate=3.75)
        assert totals["CMA CGM / APL"]["vb_importacion_usd"] == pytest.approx(450.0, rel=0.001)

    def test_one_mixed_units_flat_factura_plus_per_container_plus_per_bl(self, vb_importacion_raw):
        # Gastos Administrativos (Factura, flat) 10 + Box fee (Cntr.) 155 + SCAC (Cntr.) 42 + Doc Fee (BL, flat) 115
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=1, exchange_rate=3.75)
        assert totals["ONE"]["vb_importacion_usd"] == pytest.approx(322.0, rel=0.001)

    def test_one_cntr_concepts_scale_factura_and_bl_do_not(self, vb_importacion_raw):
        # num_containers=2: Box fee/SCAC double (155+42)*2=394; Gastos Admin (10) + Doc Fee (115) stay flat.
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=2, exchange_rate=3.75)
        assert totals["ONE"]["vb_importacion_usd"] == pytest.approx(519.0, rel=0.001)

    def test_cosco_tier_1_container(self, vb_importacion_raw):
        # SAC 120 PEN + VoBo "1 Cntr." 540 PEN, x1 container, /3.75 exchange rate.
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=1, exchange_rate=3.75)
        assert totals["COSCO / OOCL"]["vb_importacion_usd"] == pytest.approx((120.0 + 540.0) / 3.75, rel=0.001)

    def test_cosco_tier_2_to_5_containers(self, vb_importacion_raw):
        # 2 containers selects the "2 a 5 Cntr." tier (470 PEN/cntr), not the 1-cntr or 6+-cntr tiers.
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=2, exchange_rate=3.75)
        expected = (120.0 * 2 + 470.0 * 2) / 3.75
        assert totals["COSCO / OOCL"]["vb_importacion_usd"] == pytest.approx(expected, rel=0.001)

    def test_cosco_tier_6_plus_containers(self, vb_importacion_raw):
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=6, exchange_rate=3.75)
        expected = (120.0 * 6 + 390.0 * 6) / 3.75
        assert totals["COSCO / OOCL"]["vb_importacion_usd"] == pytest.approx(expected, rel=0.001)

    def test_naviera_not_on_sheet_returns_no_entry(self, vb_importacion_raw):
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=1, exchange_rate=3.75)
        assert "MAERSK / SEALAND" not in totals

    def test_totals_shape_compatible_with_get_fcl_import_local_costs(self, g_locales, mbl, vb_importacion_raw):
        totals = build_vb_importacion_totals(vb_importacion_raw, num_containers=1, exchange_rate=3.75)
        c = get_fcl_import_local_costs(g_locales, mbl, "CMA CGM / APL", vb_importacion=totals)
        assert c["vb_importacion_usd"] == pytest.approx(225.0, rel=0.001)


class TestRealImportReferenceData:
    """Real Gastos de Importacion en Callao por Naviera.xlsx data (Session
    E), transcribed via parse_g_locales_sheet/parse_mbl_sheet/
    parse_vb_importacion_sheet — hardcoded since the source file isn't
    deployed (same pattern as CONSOLIDATORS/EXPORT_NAVIERA_DATA)."""

    def test_g_locales_has_all_14_navieras(self):
        data = get_g_locales_data()
        assert len(data) == 14
        assert "MAERSK / SEALAND" in data
        assert "HAMBURG SUD / ALIANCA" in data

    def test_g_locales_cma_apl_raw_thc(self):
        # Pre-override raw sheet value (override is applied later in
        # get_fcl_import_local_costs via _NAVIERA_OVERRIDES).
        data = get_g_locales_data()
        assert data["CMA CGM / APL"]["thc_20"] == pytest.approx(65.0, rel=0.001)

    def test_mbl_has_all_navieras(self):
        data = get_mbl_data()
        assert data["MAERSK / SEALAND"]["amount"] == pytest.approx(55.0, rel=0.001)
        assert data["HYUNDAI"]["currency"] == "PEN"

    def test_vb_importacion_has_all_14_navieras_attributed(self):
        data = get_vb_importacion_data()
        assert len(data) == 14
        assert "MSC" in data
        assert "YANG MING" in data

    def test_real_data_combines_via_get_fcl_import_local_costs(self):
        g_locales = get_g_locales_data()
        mbl = get_mbl_data()
        vb_raw = get_vb_importacion_data()
        totals = build_vb_importacion_totals(vb_raw, num_containers=1, exchange_rate=3.75)
        c = get_fcl_import_local_costs(g_locales, mbl, "MAERSK / SEALAND", vb_importacion=totals)
        assert c["thc_20"] == pytest.approx(110.0, rel=0.001)
        assert c["mbl_usd"] == pytest.approx(55.0, rel=0.001)


# ── EVERGREEN import VB — Abel confirmed 2026-06-22 ──────────────────────────

class TestEvergreenImportVbAbel20260622:
    """Abel 2026-06-22: EVERGREEN import VB = $230 DO + $65 BL = $295 net.
    Previous values ($250 DO + PEN 25 gastos + $62 BL) were from the stale
    Gastos workbook; the EXPO_IMPO IMPORTACIÓN tab ($230+$65) is authoritative."""

    def test_evergreen_import_vb_net_295(self):
        totals = build_vb_importacion_totals(get_vb_importacion_data(), num_containers=1)
        assert totals["EVERGREEN"]["vb_importacion_usd"] == pytest.approx(295.0, rel=0.001)

    def test_evergreen_delivery_order_is_230_not_250(self):
        concepts = get_vb_importacion_data()["EVERGREEN"]["concepts"]
        do = next(c for c in concepts if c["concept"] == "DELIVERY ORDER")
        assert do["p_unit"] == 230.0

    def test_evergreen_bl_transmission_fee_is_65_not_62(self):
        concepts = get_vb_importacion_data()["EVERGREEN"]["concepts"]
        bl = next(c for c in concepts if "BL" in c["concept"].upper())
        assert bl["p_unit"] == 65.0
