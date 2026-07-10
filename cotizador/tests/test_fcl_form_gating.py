"""
Regression guard for the FCL form-layer gating bug (Abel F3/F4 blocker,
2026-07-02).

Bug: Session I (a9820fc) gated Section 4 coloader items for FCL on the
SERVER only (api/routes.py: extra_costeo_items = [] when mode == "fcl") —
templates/new_quote.html was never touched. With modo=FCL the form still:
  1. rendered the LCL package/dimensions/CBM block in §2,
  2. rendered §4 "Conceptos adicionales del coloader", and
  3. pre-filled Gastos Locales with LCL/coloader values (VB 90,
     Desconsolidación 25, Almacén 220, Agente de Aduana CIF 0.30/0.35
     min 70/110) via populateDefaults(), which keys on operation only.

The server guard means those values never reach a stored FCL quote — but
the form invited Abel to review/edit numbers that are silently discarded,
blocking F3/F4 validation.

These tests pin two invariants:
  A. The template gates §2's LCL CBM block and the whole §4 coloader
     section on modo=FCL — hidden AND disabled (a hidden-but-enabled
     input still submits; same bug class as the aereo consolidator
     dropdown, commit 94e29d2), with §4 defaults rebuilt on every mode
     switch so FCL→LCL→FCL cannot resurrect stale rows.
  B. The submit path ignores coloader form literals on FCL: gastos
     locales come from the engine (fcl_naviera_costs / fcl_import_costs /
     port_costs / fcl_customs_broker), never from extra_items_json.
"""

from __future__ import annotations

import json
import os
import re
import tempfile
from pathlib import Path

import pytest

_tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_tmp_db.close()
os.environ.setdefault("DB_PATH", _tmp_db.name)

from core.db import get_connection, init_db  # noqa: E402

_TEMPLATE = Path(__file__).parent.parent / "templates" / "new_quote.html"


@pytest.fixture(scope="module")
def template_src() -> str:
    return _TEMPLATE.read_text(encoding="utf-8")


# ── A. Template gating ────────────────────────────────────────────────────────

class TestSection2LclCbmBlockGated:
    """§2 package/dimensions/CBM markup must be gated behind modo != FCL."""

    def test_lcl_cbm_block_wrapper_exists(self, template_src):
        assert 'id="lcl-cbm-block"' in template_src, (
            "§2 LCL package/CBM block must be wrapped in id=\"lcl-cbm-block\" "
            "so it can be hidden for FCL"
        )

    def test_wrapper_encloses_pkg_table_add_btn_and_cbm_direct(self, template_src):
        start = template_src.index('id="lcl-cbm-block"')
        end = template_src.index("</section>", start)
        block = template_src[start:end]
        for token in ('class="pkg-table"', 'id="add-pkg-btn"', 'id="cbm-direct"',
                      'id="qty-hidden"'):
            assert token in block, f"{token} must live inside the gated lcl-cbm-block"

    def test_js_hides_block_on_fcl(self, template_src):
        assert re.search(
            r"lclCbmBlock\.style\.display\s*=\s*isFcl\s*\?\s*'none'", template_src
        ), "applyModeVisibility must hide #lcl-cbm-block when mode === 'fcl'"

    def test_js_disables_cbm_and_qty_inputs_on_fcl(self, template_src):
        # Hidden inputs still submit — volume_cbm / quantity must be disabled
        # for FCL so stale LCL dimensions can't reach an FCL quote.
        assert re.search(r"cbmDirectInp\.disabled\s*=\s*isFcl", template_src)
        assert re.search(r"qtyHiddenInp\.disabled\s*=\s*isFcl", template_src)


class TestSection4ColoaderGated:
    """§4 coloader section must be gated behind modo != FCL."""

    def test_coloader_section_has_id(self, template_src):
        assert 'id="coloader-section"' in template_src, (
            "§4 coloader section needs id=\"coloader-section\" so it can be "
            "hidden for FCL"
        )

    def test_js_hides_section_on_fcl(self, template_src):
        assert re.search(
            r"coloaderSection\.style\.display\s*=\s*isFcl\s*\?\s*'none'",
            template_src,
        ), "applyModeVisibility must hide #coloader-section when mode === 'fcl'"

    def test_js_disables_extra_items_json_on_fcl(self, template_src):
        # The dynamic §4 rows have no name attributes — only the hidden
        # extra_items_json input submits. Disable it for FCL.
        assert re.search(r"extraItemsJsonInp\.disabled\s*=\s*isFcl", template_src)

    def test_populate_defaults_guards_on_fcl(self, template_src):
        m = re.search(r"function populateDefaults\(op\)\s*\{(.*?)\n  \}",
                      template_src, re.S)
        assert m, "populateDefaults not found in new_quote.html"
        body = m.group(1)
        guard = re.search(r"if\s*\(currentMode\(\)\s*===\s*'fcl'\)", body)
        assert guard, (
            "populateDefaults must early-return for FCL — LCL coloader "
            "defaults must never be populated when modo=FCL"
        )
        # The guard must come BEFORE any default rows are added.
        first_add = min(
            (body.index(t) for t in ("addIntl(", "addLocal(") if t in body),
            default=len(body),
        )
        assert body.index("currentMode()") < first_add, (
            "the FCL guard must run before any addIntl/addLocal defaults"
        )

    def test_mode_switch_repopulates_defaults(self, template_src):
        # FCL→LCL→FCL must rebuild §4 from scratch each switch (clear-on-switch,
        # same pattern as 1cef182) — so a mode-select listener must call
        # populateDefaults.
        assert re.search(
            r"modeSel\.addEventListener\('change',.{0,200}?populateDefaults\(currentOperation\(\)\)",
            template_src, re.S,
        ), "mode-select change must re-run populateDefaults"


# ── B. Submit path: FCL gastos locales come from the engine ──────────────────

@pytest.fixture(autouse=True)
def fresh_db():
    with get_connection() as conn:
        conn.executescript("""
            DROP TABLE IF EXISTS audit_log;
            DROP TABLE IF EXISTS quotes;
            DROP TABLE IF EXISTS ref_counters;
            DROP TABLE IF EXISTS providers;
            DROP TABLE IF EXISTS provider_replies;
            DROP TABLE IF EXISTS credit_registry;
        """)
    init_db()
    yield


@pytest.fixture
def client():
    from api.app import create_app
    a = create_app()
    a.config["TESTING"] = True
    with a.test_client() as c:
        yield c


# Exactly what the (pre-fix) form's populateDefaults("importacion") serialized
# into extra_items_json — the poisoned coloader payload from Abel's screenshots.
_POISONED_COLOADER_JSON = json.dumps([
    {"concept": "Visto Bueno", "bucket": "local", "valor": 90,
     "factor": None, "min_usd": None},
    {"concept": "Desconsolidación", "bucket": "local", "valor": 25,
     "factor": None, "min_usd": None},
    {"concept": "Almacén", "bucket": "local", "valor": 220,
     "factor": None, "min_usd": None},
    {"concept": "Agente de Aduana", "bucket": "local", "cif_calc": True,
     "cif_usd": 50000, "pct_costo": 0.30, "pct_venta": 0.35,
     "min_costo": 70, "min_venta": 110,
     "valor": 150.0, "venta_neto": 175.0, "factor": None, "total": 150.0},
    {"concept": "Operative Charge", "bucket": "local", "valor": 25,
     "factor": None, "min_usd": None},
])

_FCL_IMPORT_BASE = {
    "client_name": "Gating Test SA",
    "client_email": "g@test.com",
    "mode": "fcl",
    "incoterm": "DAP",
    "operation": "importacion",
    "origin": "Shanghai, China",
    "destination": "Callao, Peru",
    "cargo_description": "cargo",
    "weight": "18000",
    "weight_unit": "kg",
    "staff_code": "GT-PC",
    "language": "es",
    "requester_type": "cliente",
    "fcl_terminal": "DPW",
    "fcl_naviera": "CMA CGM / APL",
    "fcl_container_type": "40STD",
    "num_containers": "1",
    "margin_pct": "20",
    "flete_lcl": "2000",
    "extra_items_json": _POISONED_COLOADER_JSON,
}

_LCL_ONLY_CONCEPTS = {"desconsolidación", "almacén", "operative charge"}


def _post(client, data):
    from urllib.parse import unquote
    resp = client.post("/quote/new", data=data, follow_redirects=False)
    assert resp.status_code == 302, resp.data
    ref = unquote(resp.headers["Location"].rstrip("/").split("/quote/")[-1])
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM quotes WHERE reference_code = ?", (ref,)
        ).fetchone()
    assert row is not None
    return dict(row)


class TestRestoredStateResync:
    """
    Abel F3 Item 2 (2026-07-02): the EXW screenshot (modo=FCL showing LCL
    fields, FCL selectors hidden) matches a browser that restored control
    state WITHOUT firing change events (session restore on back/refresh) or a
    cached pre-fix page — the gating itself keys on mode across all live
    paths (verified per-incoterm × client_type × operation). These pin the
    two hardening measures.
    """

    def test_form_disables_browser_autofill_restore(self, template_src):
        m = re.search(r"<form[^>]*class=\"quote-form\"[^>]*>", template_src)
        assert m and 'autocomplete="off"' in m.group(0), (
            "quote form needs autocomplete=\"off\" — browser session-restore "
            "re-applies modo/incoterm without change events, desyncing the "
            "JS-driven field gating"
        )

    def test_pageshow_reapplies_mode_visibility(self, template_src):
        assert re.search(
            r"window\.addEventListener\('pageshow',\s*applyModeVisibility\)",
            template_src,
        ), "applyModeVisibility must re-run on pageshow (restored-state resync)"

    def test_pageshow_clears_desynced_fcl_coloader_rows(self, template_src):
        m = re.search(
            r"window\.addEventListener\('pageshow',\s*function\s*\(\)\s*\{(.*?)\}\);",
            template_src, re.S,
        )
        assert m, "pageshow resync handler missing from Section 4 IIFE"
        body = m.group(1)
        assert "currentMode() === 'fcl'" in body
        assert "populateDefaults(currentOperation())" in body

    def test_new_quote_form_is_never_cached(self, client):
        resp = client.get("/quote/new")
        assert resp.status_code == 200
        assert resp.headers.get("Cache-Control") == "no-store", (
            "a cached form page keeps serving pre-fix gating after a deploy"
        )


class TestFclSubmitIgnoresColoaderLiterals:
    def test_cliente_local_gastos_locales_from_engine(self, client):
        row = _post(client, {**_FCL_IMPORT_BASE, "client_type": "cliente_local"})
        costeo = json.loads(row["costeo_json"])
        lines = json.loads(row["venta_json"])["line_items"]
        descs = {i["description"].strip().lower() for i in lines}

        # No coloader literal survives to the FCL quote
        assert not descs & _LCL_ONLY_CONCEPTS, (
            f"coloader form literals leaked into FCL venta: {descs & _LCL_ONLY_CONCEPTS}"
        )
        assert costeo.get("extra_items") is None
        # LCL consolidator VB engine path must be inert on FCL
        assert (costeo.get("visto_bueno_usd") or 0) == 0

        # Agente de Aduana comes from fcl_customs_broker, NOT the form's
        # CIF-calc venta_neto (175.0 poisoned above)
        aduana = next(i for i in lines if i["description"] == "Agente de Aduana")
        m = 1 + float(row["margin_pct"])
        expected = round(
            (costeo["fcl_customs_commission_usd"]
             + costeo["fcl_customs_gastos_operativos_usd"]) * m, 2)
        assert aduana["total"] == pytest.approx(expected)
        assert aduana["total"] != pytest.approx(175.0)

        # Naviera-doc charges present, sourced from engine fields
        thc = next(i for i in lines if i["description"] == "THC / Terminal Handling")
        assert thc["total"] == pytest.approx(round(costeo["fcl_thc_usd"] * m, 2))
        mbl = next(i for i in lines if i["description"] == "Emisión MBL")
        assert mbl["total"] == pytest.approx(round(costeo["fcl_mbl_usd"] * m, 2))

    def test_agente_internacional_registry_unaffected(self, client):
        row = _post(client, {**_FCL_IMPORT_BASE,
                             "client_type": "agente_internacional"})
        costeo = json.loads(row["costeo_json"])
        lines = json.loads(row["venta_json"])["line_items"]
        descs = {i["description"].strip().lower() for i in lines}
        assert not descs & _LCL_ONLY_CONCEPTS
        assert costeo.get("extra_items") is None
        assert costeo.get("client_type") == "agente_internacional"

    def test_lcl_control_coloader_items_still_land(self, client):
        # Same poisoned payload on LCL must still work — the guard is
        # FCL-specific, not a blanket drop of Section 4.
        data = {**_FCL_IMPORT_BASE, "mode": "lcl", "client_type": "cliente_local",
                "consolidator": "MSL", "volume_cbm": "12"}
        row = _post(client, data)
        costeo = json.loads(row["costeo_json"])
        lines = json.loads(row["venta_json"])["line_items"]
        descs = {i["description"].strip().lower() for i in lines}
        assert "almacén" in descs
        assert costeo.get("extra_items") is not None
        almacen = next(i for i in lines if i["description"] == "Almacén")
        m = 1 + float(row["margin_pct"])
        assert almacen["total"] == pytest.approx(round(220 * m, 2))


# ── C. Per-incoterm agente field gating (Abel F3/F4 2026-07-06) ───────────────
# The optional New-Quote FCL inputs must render/serialize per incoterm on the
# agente_internacional path, driven by the concept registry. FOB (Ocean Freight
# only) hides every optional input; EXW/DAP/DDP show exactly their structure.
# cliente_local is unaffected (byte-for-byte). DOM behaviour across all four
# incoterms × both client_types (incl. incoterm churn and FOB→EXW→FOB
# non-resurrection) is verified out-of-band via jsdom on the served page; these
# pins guard the server→form contract and the JS wiring.

import re as _re  # noqa: E402

from core.fcl_agente_incoterm import agente_field_visibility_map  # noqa: E402
from core.fcl_naviera_costs import export_naviera_options  # noqa: E402

# Newly gated rows/inputs and the field→row map the JS drives.
_GATED_ROW_IDS = ("row-thc-rate", "row-thc-min", "row-requires-oea-basc")
_GATED_INPUT_IDS = ("thc-rate-input", "thc-min-input", "requires-oea-basc-check")


class TestAgenteFieldsInjection:
    def test_served_page_injects_registry_map(self, client):
        # The single source of truth for the JS gating is the server-injected
        # map — it must equal agente_field_visibility_map() exactly.
        html = client.get("/quote/new").data.decode()
        m = _re.search(r"var AGENTE_FIELDS = (\{.*?\});", html)
        assert m, "AGENTE_FIELDS must be injected into the served form"
        served = json.loads(m.group(1))
        assert served == agente_field_visibility_map()

    def test_fob_hides_every_optional_field(self, client):
        html = client.get("/quote/new").data.decode()
        m = _re.search(r"var AGENTE_FIELDS = (\{.*?\});", html)
        fob = json.loads(m.group(1))["EXPO/FOB"]
        assert not any(fob.values()), (
            "FOB is Ocean-Freight-only — every optional input must be gated off"
        )

    def test_import_incoterms_show_naviera_and_thc(self, client):
        html = client.get("/quote/new").data.decode()
        fields = json.loads(_re.search(r"var AGENTE_FIELDS = (\{.*?\});", html).group(1))
        for key in ("IMPO/DAP", "IMPO/DDP"):
            assert fields[key]["naviera"] is True
            assert fields[key]["thc"] is True
        assert fields["IMPO/DDP"]["ddp_cif"] is True
        assert fields["IMPO/DAP"]["ddp_cif"] is False

    def test_exw_now_shows_naviera(self, client):
        # Abel 2026-07-10: EXW must select a naviera (VB + Gate Out come from it).
        html = client.get("/quote/new").data.decode()
        fields = json.loads(_re.search(r"var AGENTE_FIELDS = (\{.*?\});", html).group(1))
        assert fields["EXPO/EXW"]["naviera"] is True
        # FOB is still Ocean-Freight-only — no naviera.
        assert fields["EXPO/FOB"]["naviera"] is False


class TestNavieraOptionSwap:
    """The agente EXW dropdown must offer the EXPORT navieras (VB + Gate Out come
    from the EXPORTACION-CALLAO sheet); DAP/DDP and cliente_local keep the IMPORT
    list. The swap is driven by JS on the served page so it stays incoterm-aware
    without a page reload; DOM behaviour (incl. EXW↔FOB non-resurrection) is
    verified out-of-band via jsdom. These pin the server→form contract + wiring."""

    def test_export_naviera_list_injected_and_exact(self, client):
        html = client.get("/quote/new").data.decode()
        m = _re.search(r"var FCL_NAVIERAS_EXPORT = (\[.*?\]);", html)
        assert m, "FCL_NAVIERAS_EXPORT must be injected into the served form"
        served = json.loads(m.group(1))
        assert served == export_naviera_options(), (
            "EXW dropdown must offer exactly the export-VB navieras"
        )

    def test_import_naviera_list_still_injected(self, template_src):
        assert "var FCL_NAVIERAS_IMPORT =" in template_src, (
            "the import naviera list must remain for DAP/DDP + cliente_local"
        )

    def test_set_naviera_options_defined(self, template_src):
        assert "function setNavieraOptions(list)" in template_src

    def test_set_naviera_options_preserves_current_value(self, template_src):
        m = _re.search(r"function setNavieraOptions\(list\)\s*\{(.*?)\n  \}",
                       template_src, _re.S)
        assert m, "setNavieraOptions helper missing"
        body = m.group(1)
        # keeps a still-valid pick, resets otherwise
        assert "indexOf(current)" in body

    def test_swap_gated_to_agente_export(self, template_src):
        # The export list is used only when the agente field set applies AND the
        # operation is exportacion (⇒ EXW, the only EXPO incoterm with naviera).
        m = _re.search(r"function applyModeVisibility\(\)\s*\{(.*?)\n  \}",
                       template_src, _re.S)
        assert m, "applyModeVisibility not found"
        body = m.group(1)
        assert "setNavieraOptions(" in body, (
            "applyModeVisibility must swap naviera options"
        )
        assert "FCL_NAVIERAS_EXPORT" in body and "FCL_NAVIERAS_IMPORT" in body
        assert "exportacion" in body

    def test_served_export_list_covers_all_export_vb_navieras(self, client):
        # Abel's carrier must be selectable — the served list must include every
        # export-VB naviera (guards an accidental import-only render).
        html = client.get("/quote/new").data.decode()
        served = json.loads(
            _re.search(r"var FCL_NAVIERAS_EXPORT = (\[.*?\]);", html).group(1))
        for n in ("MSC", "ONE", "MAERSK", "HAPAG LLOYD", "CMA CGM", "COSCO",
                  "EVERGREEN"):
            assert n in served, n


class TestAgenteGatingTemplateWiring:
    def test_new_gated_rows_have_ids(self, template_src):
        for rid in _GATED_ROW_IDS:
            assert f'id="{rid}"' in template_src, f"{rid} row must be id-tagged"
        for iid in _GATED_INPUT_IDS:
            assert f'id="{iid}"' in template_src, f"{iid} input must be id-tagged"

    def test_apply_agente_visibility_defined_and_called(self, template_src):
        assert "function applyAgenteIncotermVisibility()" in template_src
        # applyModeVisibility must invoke it (runs last so it can hide/disable
        # rows the mode logic just showed).
        m = _re.search(r"function applyModeVisibility\(\)\s*\{(.*?)\n  \}",
                       template_src, _re.S)
        assert m and "applyAgenteIncotermVisibility()" in m.group(1), (
            "applyModeVisibility must call applyAgenteIncotermVisibility"
        )

    def test_setrow_disables_hidden_inputs(self, template_src):
        # Hidden inputs must be disabled so they can't stale-submit.
        m = _re.search(r"function setRow\(row, input, show\)\s*\{(.*?)\n  \}",
                       template_src, _re.S)
        assert m, "setRow helper missing"
        assert _re.search(r"input\.disabled\s*=\s*!show", m.group(1))

    def test_gating_restricted_to_agente_fcl(self, template_src):
        # agenteFieldSet must return null (no restriction) unless mode is FCL
        # AND client_type is agente_internacional — cliente_local unchanged.
        m = _re.search(r"function agenteFieldSet\(\)\s*\{(.*?)\n  \}",
                       template_src, _re.S)
        assert m, "agenteFieldSet helper missing"
        body = m.group(1)
        assert "modeSelect.value !== 'fcl'" in body
        assert "'agente_internacional'" in body

    def test_incoterm_operation_clienttype_rerun_visibility(self, template_src):
        # Each of these changes which agente fields apply → must re-run the
        # full visibility pass (clear-on-switch, 1cef182).
        for sel in ("incotermSel", "operationSel", "fclClientTypeSel"):
            assert _re.search(
                sel + r"\.addEventListener\('change', applyModeVisibility\)",
                template_src,
            ), f"{sel} change must re-run applyModeVisibility"

    def test_ddp_cif_inputs_disabled_when_not_ddp(self, template_src):
        # invoice/insurance are hidden for non-DDP — disable them too.
        assert _re.search(r"invoiceInput\.disabled\s*=\s*!isDdp", template_src)
        assert _re.search(r"insuranceInput\.disabled\s*=\s*!isDdp", template_src)


class TestAgenteGatingPreservesHardening:
    """The 98daf2e protections must survive the per-incoterm layer."""

    def test_form_page_still_no_store(self, client):
        assert client.get("/quote/new").headers.get("Cache-Control") == "no-store"

    def test_autocomplete_off_and_pageshow_resync_intact(self, template_src):
        m = _re.search(r"<form[^>]*class=\"quote-form\"[^>]*>", template_src)
        assert m and 'autocomplete="off"' in m.group(0)
        assert _re.search(
            r"window\.addEventListener\('pageshow', applyModeVisibility\)",
            template_src,
        ), "pageshow must still re-run applyModeVisibility (now incl. agente gating)"


# ── D. Unified agente/cliente selector (Abel 2026-07-06, "son lo mismo") ───────
# Abel confirmed in writing the agente/cliente choice repeated twice and asked to
# unify: §1 "Solicitante" (requester_type: Agente|Cliente, all modes, feeds
# SINTAD) and "Tipo de Cliente" (client_type: Cliente Local|Agente Internacional,
# the FCL pricing fork) were "lo mismo". They collapse into ONE visible selector:
#   • single control, name="client_type", values agente_internacional|cliente_local,
#     display "Agente | Cliente" — preserves the FCL pricing-fork semantics;
#   • present on EVERY mode (no longer FCL-mode-gated), since it now also feeds the
#     server-side requester_type derivation + SINTAD + display on LCL/aéreo;
#   • the separate requester_type input is REMOVED;
#   • FCL per-incoterm concept gating still keys off it via applyModeVisibility.

class TestUnifiedAgenteClienteSelector:
    def test_exactly_one_client_type_control(self, template_src):
        assert template_src.count('name="client_type"') == 1, (
            "client_type must render exactly once — the single unified selector"
        )

    def test_separate_requester_type_input_is_removed(self, template_src):
        # The old §1 "Solicitante" <select name="requester_type"> is gone — the
        # unified client_type field is the sole agente/cliente control.
        assert 'name="requester_type"' not in template_src, (
            "the separate Solicitante (requester_type) input must be removed — "
            "requester_type is now derived server-side from client_type"
        )

    def test_unified_field_offers_agente_and_cliente(self, template_src):
        # Values map to the FCL fork; display reads naturally on all modes.
        assert 'value="agente_internacional"' in template_src
        assert 'value="cliente_local"' in template_src

    def test_client_type_lives_in_section_1_top(self, template_src):
        ct = template_src.index('name="client_type"')
        s1 = template_src.index('<!-- SECTION 1')
        s2 = template_src.index('<!-- SECTION 2')
        assert s1 < ct < s2, (
            "the unified selector must live in SECTION 1 (top), alongside "
            "cliente/operación/modo/incoterm — not §3 Tarifas y Costos"
        )

    def test_client_type_precedes_costs_section(self, template_src):
        ct = template_src.index('name="client_type"')
        costs = template_src.index('<h2>3. Tarifas y Costos</h2>')
        assert ct < costs, "client_type must render before §3 Tarifas y Costos"

    def test_unified_field_present_on_all_modes_not_mode_gated(self, template_src):
        # Because it now feeds requester_type/SINTAD on LCL/aéreo, the selector
        # must NOT be hidden or disabled for non-FCL modes — the old FCL-only
        # gating on the row/select is removed.
        assert not _re.search(
            r"rowFclClientType\.style\.display\s*=\s*showOpenTransport",
            template_src,
        ), "unified selector row must NOT be FCL-mode-gated (all modes see it)"
        assert not _re.search(
            r"fclClientTypeSel\.disabled\s*=\s*!showOpenTransport",
            template_src,
        ), "unified selector must NOT be disabled for non-FCL modes"

    def test_unified_field_still_drives_fcl_incoterm_gating(self, template_src):
        # The row/select stay id-tagged and the change event still re-runs the
        # per-incoterm concept pass (FCL agente path unchanged).
        assert 'id="row-fcl-client-type"' in template_src
        assert 'id="fcl-client-type-select"' in template_src
        assert _re.search(
            r"fclClientTypeSel\.addEventListener\('change', applyModeVisibility\)",
            template_src,
        ), "the unified selector must still drive the per-incoterm gating"

    def test_client_type_serializes_single_value(self, client):
        # One field name ⇒ Werkzeug MultiDict has exactly one value; routes.py
        # f.get('client_type') is unambiguous (no last/first-write-wins).
        html = client.get('/quote/new').data.decode()
        assert html.count('name="client_type"') == 1


# ── E. SINTAD "Tipo Solicitante" derives from the unified field (all modes) ────
# The SINTAD pre-fill's "Tipo Solicitante" reads quotes.requester_type, which is
# now derived from client_type. Confirm the derived label is correct on FCL, LCL
# and aéreo — the unified field feeds SINTAD identically across modes.

_LCL_SINTAD = {
    "client_name": "SINTAD LCL SA", "client_email": "s@test.com",
    "mode": "lcl", "incoterm": "FOB", "operation": "exportacion",
    "origin": "Lima, Peru", "destination": "Hamburg, Germany",
    "cargo_description": "cargo", "weight": "500", "weight_unit": "kg",
    "volume_cbm": "2.0", "flete_lcl": "200", "consolidator": "MSL",
    "staff_code": "GT-PC", "language": "es", "margin_pct": "20",
}

_AEREO_SINTAD = {
    "client_name": "SINTAD Aereo SA", "client_email": "s@test.com",
    "mode": "aereo", "incoterm": "FOB", "operation": "exportacion",
    "origin": "Lima, Peru", "destination": "Los Angeles, USA",
    "cargo_description": "cargo", "weight": "300", "weight_unit": "kg",
    "volume_cbm": "200", "flete_rate_aereo": "5",
    "staff_code": "GT-PC", "language": "es", "margin_pct": "20",
}


def _sintad_tipo_solicitante(row: dict) -> str:
    import io
    import openpyxl
    from core.sintad_export import generate_sintad_excel
    wb = openpyxl.load_workbook(io.BytesIO(generate_sintad_excel(row)))
    ws = wb["Datos Generales"]
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Tipo Solicitante":
            return ws.cell(row=r, column=2).value
    raise AssertionError("'Tipo Solicitante' row not found in Datos Generales")


class TestSintadRequesterDerivation:
    def test_fcl_agente_sintad_tipo_solicitante_is_agente(self, client):
        row = _post(client, {**_FCL_IMPORT_BASE,
                             "client_type": "agente_internacional"})
        assert row["requester_type"] == "agente"
        assert _sintad_tipo_solicitante(row) == "Agente"

    def test_fcl_cliente_local_sintad_tipo_solicitante_is_cliente(self, client):
        row = _post(client, {**_FCL_IMPORT_BASE, "client_type": "cliente_local"})
        assert row["requester_type"] == "cliente"
        assert _sintad_tipo_solicitante(row) == "Cliente"

    def test_lcl_agente_sintad_tipo_solicitante_is_agente(self, client):
        row = _post(client, {**_LCL_SINTAD,
                             "client_type": "agente_internacional"})
        assert row["requester_type"] == "agente"
        assert _sintad_tipo_solicitante(row) == "Agente"

    def test_aereo_cliente_sintad_tipo_solicitante_is_cliente(self, client):
        row = _post(client, {**_AEREO_SINTAD, "client_type": "cliente_local"})
        assert row["requester_type"] == "cliente"
        assert _sintad_tipo_solicitante(row) == "Cliente"
